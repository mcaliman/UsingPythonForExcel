import openpyxl
from openpyxl.utils import get_column_letter

# Simple read
xlsxExample = openpyxl.load_workbook("Example.xlsx")
# Test if read only
if xlsxExample.read_only:
    print("This Excel file is read only")
else:
    print("This Excel file is not read only")

# Print all Excel file properties
print("Print all Excel file properties", xlsxExample.properties)

sheetA = xlsxExample["SheetA"]
max_row = sheetA.max_row  # get max row defined
max_column = sheetA.max_column  # get max column defined
skip_first_row = True
print("max_row:", max_row, "max_column", max_column)
first_row = 1
if skip_first_row:
    first_row = 2
for r in range(first_row, max_row + 1):
    # Col A
    addrA = "A" + str(r)
    id = sheetA[addrA].value
    # Col B
    addrB = "B" + str(r)
    date = sheetA[addrB].value
    # Col C
    addrC = "C" + str(r)
    amount = sheetA[addrC].value
    print("row:", r, id, date, amount)
print("\n")

# scan all cells - scan programmatically
for r in range(1, max_row + 1):
    for c in range(1, max_column + 1):
        cell = sheetA.cell(row=r, column=c)  # get the value of row r and column c
        letter = get_column_letter(c)  # get the column letter (i.e. A,B,..AA...)
        print("", letter + str(r), "=", cell.value, end=" ")
    print()

